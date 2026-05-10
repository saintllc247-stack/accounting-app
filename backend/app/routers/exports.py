import io
import csv
from datetime import date

from datetime import datetime

from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User, Transaction, Client, Invoice, InvoiceItem, Category
from app.auth import get_current_user

router = APIRouter(prefix="/api/exports", tags=["exports"])


@router.get("/transactions/excel")
def export_transactions_excel(
    from_date: date = Query(None, alias="from"),
    to_date: date = Query(None, alias="to"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from openpyxl import Workbook

    q = db.query(Transaction).filter(Transaction.user_id == user.id)
    if from_date:
        q = q.filter(Transaction.date >= from_date)
    if to_date:
        q = q.filter(Transaction.date <= to_date)
    txns = q.order_by(Transaction.date.desc()).all()
    cats = {c.id: c.name for c in db.query(Category).filter(Category.user_id == user.id).all()}
    clients = {c.id: c.name for c in db.query(Client).filter(Client.user_id == user.id).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    ws.append(["Дата", "Тип", "Категория", "Клиент", "Описание", "Сумма"])
    for t in txns:
        ws.append([t.date, "Доход" if t.type == "income" else "Расход", cats.get(t.category_id, ""), clients.get(t.client_id, ""), t.description, t.amount])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions.xlsx"})


@router.get("/clients/excel")
def export_clients_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from openpyxl import Workbook

    clients = db.query(Client).filter(Client.user_id == user.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["Название", "Email", "Телефон", "Адрес", "ИНН"])
    for c in clients:
        ws.append([c.name, c.email, c.phone, c.address, c.tin])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=clients.xlsx"})


@router.get("/invoices/excel")
def export_invoices_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from openpyxl import Workbook

    invoices = db.query(Invoice).filter(Invoice.user_id == user.id).order_by(Invoice.created_at.desc()).all()
    clients = {c.id: c.name for c in db.query(Client).filter(Client.user_id == user.id).all()}
    status_map = {"draft": "Черновик", "sent": "Отправлен", "paid": "Оплачен", "cancelled": "Отменён"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Счета"
    ws.append(["№ счёта", "Клиент", "Дата", "Статус", "Сумма"])
    for inv in invoices:
        ws.append([inv.invoice_number, clients.get(inv.client_id, ""), str(inv.issue_date), status_map.get(inv.status, inv.status), inv.total_amount])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=invoices.xlsx"})


@router.get("/invoices/{inv_id}/xlsx")
def export_invoice_xlsx(inv_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    inv = db.query(Invoice).filter(Invoice.id == inv_id, Invoice.user_id == user.id).first()
    if not inv:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Invoice not found")

    client = db.query(Client).filter(Client.id == inv.client_id).first()
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).all()
    status_map = {"draft": "Черновик", "sent": "Отправлен", "paid": "Оплачен", "cancelled": "Отменён"}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Счёт {inv.invoice_number}"
    ws.page_setup.orientation = "portrait"

    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = f"СЧЁТ № {inv.invoice_number}"
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Статус: {status_map.get(inv.status, inv.status)}"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["От:"]), ws.merge_cells("A4:E4")
    ws["A4"].value = user.company_name or user.username
    ws.append(["Email:", user.email])
    ws.append([])

    if client:
        ws.append(["Кому:"])
        ws.merge_cells("A7:E7")
        ws["A7"].value = client.name
        if client.email:
            ws.append(["Email:", client.email])
        if client.phone:
            ws.append(["Телефон:", client.phone])
        if client.address:
            ws.append(["Адрес:", client.address])
        if client.tin:
            ws.append(["ИНН:", client.tin])

    ws.append([])
    ws.append(["Дата выписки:", str(inv.issue_date), "", "Срок оплаты:", str(inv.due_date)])

    ws.append([])
    headers = ["Описание", "Кол-во", "Цена", "Сумма"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style="thin"))

    for item in items:
        ws.append([item.description, item.quantity, item.unit_price, item.amount])

    ws.append([])
    ws.append(["", "", "ИТОГО:", inv.total_amount])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    thick = Border(top=Side(style="double"))
    for cell in ws[ws.max_row]:
        cell.border = thick

    if inv.notes:
        ws.append([])
        ws.append(["Примечание:", inv.notes])

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=invoice_{inv.invoice_number}.xlsx"})


def parse_date_flexible(val: str) -> date:
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%Y %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return date.today()


@router.post("/transactions/import")
def import_transactions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from datetime import datetime
    from io import BytesIO

    imported = 0
    rows = []

    if file.filename and file.filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file.file.read()), read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        content = file.file.read().decode("utf-8-sig")
        lines = content.splitlines()
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(lines[0] if len(lines) > 1 else lines[0] + "\n" + lines[0])
        rows = list(csv.DictReader(lines, dialect=dialect))

    if not rows:
        return {"imported": 0}

    headers = set(rows[0].keys())

    marketplace = "orderId" in headers or "order amount" in headers

    for row in rows:
        if marketplace:
            try:
                raw = str(row.get("order amount", "0")).replace(",", ".").replace(" ", "")
                amt = float(raw)
            except ValueError:
                continue
            date_str = row.get("acceptedDate") or row.get("createdDate") or ""
            txn = Transaction(
                user_id=user.id,
                type="income",
                amount=amt,
                description=f"Заказ #{row.get('orderId', '')} — {row.get('customer', '').strip()}",
                date=parse_date_flexible(date_str),
            )
        else:
            try:
                raw = str(row.get("amount", "0")).replace(",", ".").replace(" ", "")
                amt = float(raw)
            except ValueError:
                continue
            txn = Transaction(
                user_id=user.id,
                type=row.get("type", "income"),
                amount=amt,
                description=row.get("description", ""),
                date=parse_date_flexible(row.get("date", "")),
            )
        db.add(txn)
        imported += 1
    db.commit()
    return {"imported": imported}
